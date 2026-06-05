import os
import sys
import csv
from pathlib import Path

from openpyxl import load_workbook, Workbook

MAX_ROWS = 1_000_000
PROGRESS_INTERVAL = 100_000


def get_base_folder():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def ask_delimiter(default=';'):
    prompt = (
        f"Auto-detected delimiter is '{default}'. If this is correct, press Enter. "
        "If not, enter the correct delimiter (comma, semicolon, tab, pipe or single char): "
    )
    print()
    value = input(prompt).strip()
    if not value:
        return default
    if value.lower() in ('comma', ','):
        return ','
    if value.lower() in ('semicolon', ';'):
        return ';'
    if value.lower() in ('tab', '\t'):
        return '\t'
    if value.lower() in ('pipe', '|'):
        return '|'
    return value[0]


def ask_max_rows(default=MAX_ROWS, maximum=MAX_ROWS):
    prompt = (
        f"Enter max rows per sheet (1-{maximum:,}), or press Enter to keep default [{default:,}]: "
    )
    print()
    while True:
        value = input(prompt).strip().replace(',', '')
        if not value:
            return default
        if not value.isdigit():
            print(f"Please enter a whole number between 1 and {maximum:,}.")
            continue
        rows = int(value)
        if 1 <= rows <= maximum:
            return rows
        print(f"Value must be between 1 and {maximum:,}.")


def find_input_file(folder):
    candidates = [
        f for f in os.listdir(folder)
        if (f.lower().endswith(".xlsx") or f.lower().endswith(".csv"))
        and not f.startswith("~$")
        and "_SPLIT" not in f
    ]

    if not candidates:
        raise FileNotFoundError("No .xlsx or .csv file found.")

    if len(candidates) == 1:
        return os.path.join(folder, candidates[0])

    print("\nMultiple files found:")
    for i, file in enumerate(candidates, start=1):
        print(f"{i}. {file}")

    choice = int(input("\nSelect file number: "))
    return os.path.join(folder, candidates[choice - 1])


def process_sheet(source_ws, target_wb, max_rows):

    print(f"\nProcessing sheet: {source_ws.title}")

    row_iterator = source_ws.iter_rows(values_only=True)

    try:
        header = next(row_iterator)
    except StopIteration:
        print("Sheet is empty. Skipped.")
        return

    part_no = 1
    current_row_count = 0
    total_processed = 0

    sheet_name = f"{source_ws.title[:20]}_P{part_no}"
    target_ws = target_wb.create_sheet(title=sheet_name)

    target_ws.append(header)

    print(f"Creating: {sheet_name}")

    for row in row_iterator:

        if current_row_count >= max_rows:

            part_no += 1

            sheet_name = f"{source_ws.title[:20]}_P{part_no}"

            target_ws = target_wb.create_sheet(title=sheet_name)
            target_ws.append(header)

            current_row_count = 0

            print(f"Creating: {sheet_name}")

        target_ws.append(row)

        current_row_count += 1
        total_processed += 1

        if total_processed % PROGRESS_INTERVAL == 0:
            print(f"  {total_processed:,} rows processed")

    print(
        f"Completed sheet '{source_ws.title}' "
        f"({total_processed:,} data rows)"
    )


def process_csv(input_file, output_file, delimiter, max_rows):

    print(f"\nProcessing CSV: {os.path.basename(input_file)}")
    print(f"Using delimiter: '{delimiter}'")

    base = Path(input_file).stem
    part_no = 1
    total_processed = 0

    # Create write-only workbook for single XLSX output with multiple sheets
    target_wb = Workbook(write_only=True)

    # Open input with a large buffer for speed and tolerant encoding
    with open(input_file, newline='', encoding='utf-8', errors='replace', buffering=1024*1024) as inf:
        reader = csv.reader(inf, delimiter=delimiter, quotechar='"')

        try:
            header = next(reader)
        except StopIteration:
            print("CSV is empty. Skipped.")
            return

        sheet_name = f"{base[:20]}_P{part_no}"
        target_ws = target_wb.create_sheet(title=sheet_name)
        target_ws.append(header)

        print(f"Creating: {sheet_name}")

        current_row_count = 0

        for row in reader:

            if current_row_count >= max_rows:
                part_no += 1
                sheet_name = f"{base[:20]}_P{part_no}"
                target_ws = target_wb.create_sheet(title=sheet_name)
                target_ws.append(header)
                current_row_count = 0
                print(f"Creating: {sheet_name}")

            target_ws.append(row)

            current_row_count += 1
            total_processed += 1

            if total_processed % PROGRESS_INTERVAL == 0:
                print(f"  {total_processed:,} rows processed")

    # Save single XLSX with multiple sheets
    print("\nSaving output file...")
    target_wb.save(output_file)

    print(f"Completed CSV -> XLSX ({total_processed:,} data rows)")


def main():

    print("=" * 60)
    print("CSV and Excel File Splitter")
    print(f"Max rows per sheet: {MAX_ROWS:,}")
    print("=" * 60)

    try:

        folder = get_base_folder()

        input_file = find_input_file(folder)

        print(f"\nInput file:")
        print(os.path.basename(input_file))

        ext = Path(input_file).suffix.lower()

        if ext == '.csv':
            output_file = str(
                Path(input_file).with_name(
                    Path(input_file).stem + "_SPLIT.xlsx"
                )
            )

            # Detect delimiter, then allow user override
            with open(input_file, newline='', encoding='utf-8', errors='replace', buffering=1024*1024) as inf:
                sample = inf.read(8192)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=';,|\t')
                default_delimiter = dialect.delimiter
            except csv.Error:
                default_delimiter = ';' if sample.count(';') > sample.count(',') else ','

            delimiter = ask_delimiter(default_delimiter)
            max_rows = ask_max_rows()
            process_csv(input_file, output_file, delimiter, max_rows)
            print("\n" + "=" * 60)
            print("COMPLETED")
            print(output_file)
            print("=" * 60)
        else:
            # else assume xlsx
            output_file = str(
                Path(input_file).with_name(
                    Path(input_file).stem + "_SPLIT.xlsx"
                )
            )

            print("\nOpening workbook...")

            source_wb = load_workbook(
                input_file,
                read_only=True,
                data_only=True
            )

            # Use write-only workbook for much faster writes and lower memory
            target_wb = Workbook(write_only=True)

            print(
                f"\nFound {len(source_wb.sheetnames)} sheet(s):"
            )

            for sheet in source_wb.sheetnames:
                print(f" - {sheet}")

            max_rows = ask_max_rows()
            for sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]
                process_sheet(source_ws, target_wb, max_rows)

            print("\nSaving output file...")

            target_wb.save(output_file)

            source_wb.close()

            print("\n" + "=" * 60)
            print("COMPLETED")
            print(output_file)
            print("=" * 60)

    except Exception as e:
        print("\nERROR:")
        print(str(e))

    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()